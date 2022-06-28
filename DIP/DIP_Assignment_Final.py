import numpy as np
import cv2 
from matplotlib import pyplot as pt
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import pytesseract 

# Create Excel file for text extraction
wb = Workbook()

# Function: Determine the number of index to subtract from the total contour
def sub_index(n, contours, img):
    # Create mask with the img size 
    img_mask = np.zeros_like(img)
    # Draw last index of filled contour on mask
    cnt_img = cv2.drawContours(img_mask ,contours, len(contours) - n, (255,255,255), -1)
    # Count non zero pixels (white) to determine the size of the region of interest
    count_cnt = np.count_nonzero(cnt_img)
    repeat = True
    # Drawings will take place in a large area, 
    # thus determining the most non zero pixel values in the region
    # If the region is small, will switch to the next contour index by subtracting n 
    while repeat:
        if count_cnt <= 5000000:
            n += 1
            cnt_img = cv2.drawContours(img_mask, contours, len(contours) - n, (255,255,255), -1)
            count_cnt = np.count_nonzero(cnt_img)
        elif count_cnt >= 5000000:
            repeat = False
            
    # return n to be subtracted to select the index of the contour
    return n

def extract_info(num, img, cnts):
    # Open active excel file 
    ws = wb.active
    ws.title = "Drawing Info"
    
    # Assigning the information in fields 
    ws['A1'].value = "IMAGE NO."
    ws['B1'].value = "TITLE:"
    ws['C1'].value = "DRAWING NO.:"
    ws['D1'].value = "CONTRACTOR:"
    ws['E1'].value = "DRAWN BY:"
    ws['F1'].value = "CHECKED BY:"
    ws['G1'].value = "APPROVED BY:"
    ws['H1'].value = "STATUS:"
    ws['I1'].value = "PROJECT NO:"
    
    # Adjusting column width to each field
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 17
    
    # Loop image number
    count = 1
    for x in range(21):
        count += 1
        field = "A" + str(count)
        ws[field].value = x + 1
    
    # Invert color contour to cover the drawing instead of the information 
    inverted_cnts = cv2.threshold(cnts,0,255,cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    # Bitwise AND the image and the contour to only display the information 
    output = cv2.bitwise_and(img, inverted_cnts)
    
    # Create a a mask with the same size of the image 
    result = np.ones_like(img)
    # Convert it to white color 
    result[:,:] = 255
    result = result.astype(np.uint8)
    # Apply only the information on a white image 
    result[inverted_cnts == 255] = output[inverted_cnts == 255]
    
    
    # Remove horizontal lines
    result_copy = result.copy()
    result_thresh = cv2.threshold(result, 0,255,cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU) [1]

    #Kernel
    horizontal_kernel= cv2.getStructuringElement(cv2.MORPH_RECT, (40,1))
    remove_horizontal = cv2.morphologyEx(result_thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
    cnts_line = cv2.findContours(remove_horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts_line = cnts_line[0] if len(cnts_line) == 2 else cnts_line[1]
    for c in cnts_line:
        cv2.drawContours(result_copy, [c], -1, (255,255,255),5)
    
    # Remove vertical lines
    #Kernel
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1,40))
    remove_vertical = cv2.morphologyEx(result_thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
    cnts_line = cv2.findContours(remove_vertical, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts_line = cnts_line[0] if len(cnts_line) == 2 else cnts_line[1]
    for c in cnts_line:
        cv2.drawContours(result_copy, [c], -1, (255,255,255),5)
        
        
    # Read text
    pytesseract.pytesseract.tesseract_cmd=r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    text = pytesseract.image_to_string(result_copy, lang='eng', config='--psm 3')
    # print(text)
    tokens = text.split()
    for n, i in enumerate(tokens):
        if 'DRAWING' in i:
            join_drawing_number = ' '.join(tokens[n: n + 2])
            if join_drawing_number == 'DRAWING NUMBER:' or join_drawing_number == 'DRAWING NO:' or join_drawing_number == 'DRAWING NO.:':
                drawing_number = ' '.join(tokens[n + 2: n + 3])
                if len(drawing_number) <= 2:
                    drawing_number = '-'.join(tokens[n + 2: n + 10])
                drawing_num_field = "C" + str(num + 1)
                ws[drawing_num_field].value = drawing_number
                
                
        if 'TITLE' in i:
            join_title = ' '.join(tokens[n: n + 3])
            rev = join_title[len(join_title):13:-1]
        
            if rev[::-1] == 'DESIGN':
                drawing_title = ' '.join(tokens[n + 1: n + 3])
                drawing_title_field = "B" + str(num + 1)
                ws[drawing_title_field].value = drawing_title
            else:
                drawing_title = ' '.join(tokens[n + 2: n + 3])
                if drawing_title == 'DRAWING':
                    drawing_title = ' '.join(tokens[n + 1: n + 2])
                    drawing_title_field = "B" + str(num + 1)
                    ws[drawing_title_field].value = drawing_title
                else:
                    drawing_title = ' '.join(tokens[n + 1: n + 3])
                    drawing_title_field = "B" + str(num + 1)
                    ws[drawing_title_field].value = drawing_title
                
                
        if 'DRAWN' in i:
            join_drawn_by = ' '.join(tokens[n: n + 2])
            if join_drawn_by == 'DRAWN BY:' or join_drawn_by == 'DRAWN BY':
                drawn_by = ' '.join(tokens[n + 2: n + 3])
            else:
                drawn_by = ' '.join(tokens[n + 1: n + 2])
            drawn_field = "E" + str(num + 1)
            ws[drawn_field].value = drawn_by
            
        if 'CHECKED' in i:
            join_checked = ' '.join(tokens[n : n + 2])
            if join_checked == 'CHECKED BY:' or join_checked == 'CHECKED BY':
                checked = ' '.join(tokens[n + 2: n + 3])
            else:
                checked = ' '.join(tokens[n + 1: n + 2])
            checked_field = "F" + str(num + 1)
            ws[checked_field].value = checked
            
        if 'APPROVED' in i:
            join_approved = ' '.join(tokens[n: n + 2])
            if join_approved == 'APPROVED BY:' or join_approved == 'APPROVED BY':
                approved = ' '.join(tokens[n + 2: n + 3])
            else:
                approved = ' '.join(tokens[n + 1: n + 2])
            approve_field = "G" + str(num + 1)
            ws[approve_field].value = approved
            
        if 'CONTRACTOR' in i:
            join_contractor = ' '.join(tokens[n : n + 1])
            if join_contractor == 'CONTRACTOR:' or join_contractor == 'CONTRACTOR':
                contractor = ' '.join(tokens[n + 1: n + 2])
            else:
                contractor = 'None'
            contractor_field = "D" + str(num + 1)
            ws[contractor_field].value = contractor
            
        if 'PROJECT' in i:
            join_project = ' '.join(tokens[n: n + 2])
            if join_project == 'PROJECT NO:' or join_project == 'PROJECT NO':
                project = ' '.join(tokens[n + 3: n + 4])
                if project == 'SU':
                    project = '-'.join(tokens[n + 3: n + 6])
                elif project == 'CAD':
                    project = '-'.join(tokens[n + 10: n + 13])
                elif project == 'PRJ':
                    project = '-'.join(tokens[n + 2: n + 5])
            else:
                project = 'None'
            project_field = "I" + str(num + 1)
            ws[project_field].value = project    
            
        if 'STATUS' in i:
            join_status = ' '.join(tokens[n: n + 1])
            if join_status == 'STATUS:' or join_status == 'STATUS':
                status = ' '.join(tokens[n + 2: n + 3])
                if status == 'PROJECT':
                    status = ''.join(tokens[n + 1])
                if status == 'PAGE' or status == 'PAGE:':
                    status = ''.join(tokens[n + 1])
                    if status == 'PAGE:':
                        status = ''.join(tokens[n + 10])
                status_field = "H" + str(num + 1)
                ws[status_field].value = status
            
            
    
    # Save the excel file 
    wb.save('Information.xlsx')
    print("Information " + str(num) + " is saved. ")


# Extract all 21 images, image 21 will be randomize by the lecturer
def extract_image():
    # Image number 
     number = 0
     run = True
     # Run all images with while loop
     while run:
         number += 1
         if number < 10:
             image_file = "0" + str(number) + ".png"
         elif number >= 10:
             image_file = str(number) + ".png"
             
         # Read image
         image = cv2.imread(image_file, 0)
         # Inverted image 
         inverted_image = cv2.threshold(image,0,255,cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
         # Blur image for better edge detection
         blur = cv2.GaussianBlur(inverted_image, (5,5), cv2.BORDER_DEFAULT)
         # Invert to original color image
         img = cv2.threshold(blur,0,255,cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
         # Find contours in image 
         contours, hierarchy = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
         
         # Create mask 
         mask = np.zeros_like(img)
         # Find the subtracting number to find the drawing contour index
         x = sub_index(1, contours, img)
        
         # Draw selected contour on mask 
         cnts = cv2.drawContours(mask, contours, len(contours) - x , (255,255,255), -1)
         
         # Bitwise AND the mask and the image to cover the information
         output = cv2.bitwise_and(img, mask)
         # Create a white mask
         result = np.ones_like(img)
         result[:,:] = 255
         # Convert data type to uint8
         result = result.astype(np.uint8)
         # Implement the drawing from output into result variable (White mask)
         result[cnts==255] = output[cnts == 255]
        
         # Save extracted drawings as new .png file 
         cv2.imwrite("{}extracted.png".format(number), result)
         print("Image " + str(number) + " is saved. ")
         
         # Extract Information from the image
         extract_info(number, img, cnts)
         
         # End loop after all 21 images are extracted
         if number == 21:
            run = False

# Start program
extract_image()

